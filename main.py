import time
from collections import defaultdict
from fastapi import FastAPI, Depends, Cookie, Response, Request, HTTPException, status
from fastapi.responses import HTMLResponse, RedirectResponse, StreamingResponse
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel, validator
from sqlalchemy.orm import Session
from sqlalchemy import func
from database import init_db, get_db, Producto, Venta, Cliente, Gasto, Historial, Usuario
from datetime import datetime, timedelta, timezone
from auth import verificar_password, crear_token, verificar_token, ADMIN_USER, ADMIN_PASSWORD
from auth import pwd_context, generar_reset_token, enviar_correo, ADMIN_EMAIL, reset_tokens
from typing import Optional, List
from fastapi.staticfiles import StaticFiles
import io
import urllib.parse
import re
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from dotenv import load_dotenv
from scheduler import iniciar_scheduler
from whatsapp_service import whatsapp_service
import os
import html

load_dotenv()

def sanitize_html(text: str) -> str:
    return html.escape(text).strip()

def sanitize_filename(text: str) -> str:
    text = re.sub(r'[^a-zA-Z0-9_\-\.]', '', text)
    return text[:100]

# --- HELPERS DE AUTENTICACIÓN ---
def get_current_user(token: str = Cookie(default=None)):
    if not token:
        return None
    email = verificar_token(token)
    if not email:
        return None
    db = next(get_db())
    return db.query(Usuario).filter(Usuario.email == email, Usuario.activo == True).first()

def require_auth(token: str = Cookie(default=None)):
    user = get_current_user(token)
    if not user:
        raise HTTPException(status_code=401, detail="No autorizado")
    return user

def require_rol(token: str = Cookie(default=None), roles: List[str] = ["admin"]):
    user = get_current_user(token)
    if not user or user.rol not in roles:
        raise HTTPException(status_code=403, detail="No autorizado")
    return user

# --- RATE LIMITING ---
login_attempts = defaultdict(list)
MAX_ATTEMPTS = 5
LOCKOUT_TIME = 300

def check_rate_limit(ip: str) -> bool:
    ahora = time.time()
    login_attempts[ip] = [t for t in login_attempts[ip] if ahora - t < LOCKOUT_TIME]
    if len(login_attempts[ip]) >= MAX_ATTEMPTS:
        return False
    login_attempts[ip].append(ahora)
    return True

def sanitize_input(value: str) -> str:
    if not value:
        return value
    return re.sub(r'[<>"\']', '', value)

# Cargar variables de entorno
load_dotenv()

app = FastAPI(
    title="CarniMarket API",
    docs_url=None,
    redoc_url=None
)

app.add_middleware(
    CORSMiddleware,
    allow_origins=["https://carnimarket-production.up.railway.app"],
    allow_credentials=True,
    allow_methods=["GET", "POST", "PUT", "DELETE"],
    allow_headers=["*"],
)

# Configuración de archivos estáticos para CSS, JS e Imágenes
app.mount("/static", StaticFiles(directory="static"), name="static")

# --- UTILIDADES DE TIEMPO (ZONA HORARIA COLOMBIA UTC-5) ---
def obtener_hora_colombia():
    return datetime.now(timezone.utc).replace(tzinfo=None) - timedelta(hours=5)

@app.get("/test-db")
def test_db(db: Session = Depends(get_db)):
    if os.getenv("DEBUG") != "true":
        raise HTTPException(status_code=404)
    from auth import pwd_context
    count = db.query(Usuario).count()
    usuario = db.query(Usuario).filter(Usuario.email == "stevenhm03@gmail.com").first()
    
    if usuario:
        test_pass = verificar_password("C@rniMarket", usuario.password_hash)
        return {"usuarios": count, "admin_existe": True, "password_valida": test_pass, "rol": usuario.rol, "activo": usuario.activo}
    else:
        nuevo = Usuario(
            email="stevenhm03@gmail.com",
            password_hash=pwd_context.hash("C@rniMarket"),
            nombre="Administrador",
            rol="admin",
            activo=True
        )
        db.add(nuevo)
        db.commit()
        return {"usuarios": count + 1, "admin_creado": True}

@app.on_event("startup")
def startup():
    init_db()
    try:
        iniciar_scheduler()
    except Exception:
        pass

# --- MODELOS DE ENTRADA (PYDANTIC) ---
class LoginRequest(BaseModel):
    email: str
    password: str
    
    @validator('email', 'password')
    def validate_length(cls, v):
        if len(v) < 3:
            raise ValueError('Mínimo 3 caracteres')
        if len(v) > 50:
            raise ValueError('Máximo 50 caracteres')
        return v

class ProductoRequest(BaseModel):
    nombre: str
    stock: float
    minimo: float
    precio_kilo: float
    tipo: str = "kilo"

class ProductoUpdate(BaseModel):
    nombre: Optional[str] = None
    stock: Optional[float] = None
    minimo: Optional[float] = None
    precio_kilo: Optional[float] = None
    tipo: Optional[str] = None

class VentaRequest(BaseModel):
    producto: str
    cantidad: float  # kilos, gramos (se convierte a kilos), o platos
    unidad: str = "kilo"  # "kilo", "gramos", "plato"
    cliente_nombre: str = "Cliente general"
    pagado: str = "encargado"  #默认为encargado
    fecha_venta: Optional[str] = None
    fecha_vencimiento: Optional[str] = None
    notas: str = ""
    
    @validator('producto', 'cliente_nombre', 'notas')
    def sanitize_strings(cls, v):
        return sanitize_input(v)
    
    @validator('cantidad')
    def validate_cantidad(cls, v):
        if v <= 0 or v > 1000:
            raise ValueError('Cantidad inválida')
        return v

class ClienteRequest(BaseModel):
    nombre: str
    telefono: str = ""
    direccion: str = ""
    
    @validator('nombre', 'telefono', 'direccion')
    def sanitize_strings(cls, v):
        return sanitize_input(v)

class ClienteUpdate(BaseModel):
    nombre: Optional[str] = None
    telefono: Optional[str] = None
    direccion: Optional[str] = None

class GastoRequest(BaseModel):
    descripcion: str
    categoria: str
    monto: float

# --- MODELO DE REGISTRO DE CAMBIOS ---
class RegistroCambio(BaseModel):
    tipo: str
    descripcion: str
    fecha: datetime

# --- RUTAS DE NAVEGACIÓN HTML ---

@app.get("/", response_class=HTMLResponse)
def inicio():
    with open("templates/index.html", "r", encoding="utf-8") as f:
        return f.read()

@app.get("/login", response_class=HTMLResponse)
def login_page():
    with open("templates/login.html", "r", encoding="utf-8") as f:
        return f.read()

@app.get("/admin", response_class=HTMLResponse)
def admin(token: str = Cookie(default=None)):
    if not token or not verificar_token(token):
        return RedirectResponse(url="/login", status_code=302)
    with open("templates/admin.html", "r", encoding="utf-8") as f:
        return f.read()

# --- AUTENTICACIÓN ---

@app.post("/auth/login")
def auth_login(data: LoginRequest, response: Response, request: Request):
    client_ip = request.client.host if request.client else "unknown"
    
    if not check_rate_limit(client_ip):
        return {"error": "Demasiados intentos. Intenta más tarde."}
    
    db = next(get_db())
    usuario = db.query(Usuario).filter(Usuario.email == data.email.lower().strip()).first()
    
    if not usuario or not usuario.activo:
        login_attempts[client_ip].append(time.time())
        return {"error": "Credenciales incorrectas"}
    
    if not verificar_password(data.password, usuario.password_hash):
        login_attempts[client_ip].append(time.time())
        return {"error": "Credenciales incorrectas"}
    
    usuario.ultimo_login = datetime.now()
    db.commit()
    
    token = crear_token({"sub": usuario.email})
    login_attempts[client_ip] = []
    
    response.set_cookie(
        key="token", 
        value=token, 
        httponly=True, 
        max_age=28800, 
        samesite="lax", 
        secure=True
    )
    return {"token": token, "rol": usuario.rol, "nombre": usuario.nombre}

@app.post("/auth/logout")
def logout(response: Response):
    response.delete_cookie("token")
    return {"mensaje": "Sesión cerrada"}

@app.get("/auth/verificar")
def verificar_auth(token: str = Cookie(default=None)):
    try:
        if not token:
            return {"rol": "empleado", "nombre": ""}
        
        email = verificar_token(token)
        if not email:
            raise HTTPException(status_code=401, detail="Token invalido")
        
        db = next(get_db())
        usuario = db.query(Usuario).filter(Usuario.email == email).first()
        if not usuario:
            return {"rol": "empleado", "nombre": ""}
        
        return {"rol": usuario.rol or "empleado", "nombre": usuario.nombre or ""}
    except HTTPException:
        raise
    except Exception:
        return {"rol": "empleado", "nombre": ""}

# --- GESTIÓN DE USUARIOS ---
@app.get("/admin/usuarios")
def listar_usuarios(db: Session = Depends(get_db), usuario: str = Cookie(default=None)):
    if not verificar_token(usuario):
        raise HTTPException(status_code=401, detail="No autorizado")
    
    email = verificar_token(usuario)
    if not email:
        raise HTTPException(status_code=401, detail="Token invalido")
    
    u = db.query(Usuario).filter(Usuario.email == email).first()
    if not u or u.rol not in ["admin"]:
        raise HTTPException(status_code=403, detail="Solo admins")
    
    usuarios = db.query(Usuario).order_by(Usuario.fecha_registro.desc()).all()
    resultado = []
    for x in usuarios:
        try:
            ultimo = ""
            if x.ultimo_login is not None:
                ultimo = x.ultimo_login.strftime("%Y-%m-%d %H:%M")
            resultado.append({
                "id": x.id, 
                "email": x.email or "", 
                "nombre": x.nombre or "", 
                "rol": x.rol or "empleado", 
                "activo": x.activo or False,
                "ultimo_login": ultimo
            })
        except Exception:
            resultado.append({
                "id": x.id, 
                "email": "", 
                "nombre": "", 
                "rol": "empleado", 
                "activo": False, 
                "ultimo_login": ""
            })
    return resultado

@app.post("/admin/usuario")
def crear_usuario(data: dict, db: Session = Depends(get_db), token: str = Cookie(default=None)):
    if not verificar_token(token):
        raise HTTPException(status_code=401, detail="No autorizado")
    u = db.query(Usuario).filter(Usuario.email == verificar_token(token)).first()
    if not u or u.rol not in ["admin"]:
        raise HTTPException(status_code=403, detail="Solo admins")
    
    if db.query(Usuario).filter(Usuario.email == data.get("email", "")).first():
        return {"error": "Email ya existe"}
    
    nuevo = Usuario(
        email=data.get("email", "").lower().strip(),
        password_hash=pwd_context.hash(data.get("password", "")),
        nombre=data.get("nombre", "").strip(),
        rol=data.get("rol", "empleado")
    )
    db.add(nuevo)
    db.commit()
    return {"mensaje": "Usuario creado"}

@app.put("/admin/usuario/{id}")
def editar_usuario(id: int, data: dict, db: Session = Depends(get_db), token: str = Cookie(default=None)):
    if not verificar_token(token):
        raise HTTPException(status_code=401, detail="No autorizado")
    u = db.query(Usuario).filter(Usuario.email == verificar_token(token)).first()
    if not u or u.rol not in ["admin"]:
        raise HTTPException(status_code=403, detail="Solo admins")
    
    usuario = db.query(Usuario).filter(Usuario.id == id).first()
    if not usuario:
        return {"error": "Usuario no existe"}
    
    if data.get("nombre"):
        usuario.nombre = data["nombre"]
    if data.get("rol"):
        usuario.rol = data["rol"]
    if data.get("activo") is not None:
        usuario.activo = data["activo"]
    if data.get("password"):
        usuario.password_hash = pwd_context.hash(data["password"])
    
    db.commit()
    return {"mensaje": "Usuario actualizado"}

@app.delete("/admin/usuario/{id}")
def eliminar_usuario(id: int, db: Session = Depends(get_db), token: str = Cookie(default=None)):
    if not verificar_token(token):
        raise HTTPException(status_code=401, detail="No autorizado")
    u = db.query(Usuario).filter(Usuario.email == verificar_token(token)).first()
    if not u or u.rol not in ["admin"]:
        raise HTTPException(status_code=403, detail="Solo admins")
    
    usuario = db.query(Usuario).filter(Usuario.id == id).first()
    if not usuario:
        return {"error": "Usuario no existe"}
    
    db.delete(usuario)
    db.commit()
    return {"mensaje": "Usuario eliminado"}

@app.post("/auth/recuperar")
def solicitar_recuperacion(data: dict, db: Session = Depends(get_db)):
    email = data.get("email", "").lower().strip()
    
    usuario = db.query(Usuario).filter(Usuario.email == email).first()
    if usuario:
        import secrets
        token = secrets.token_urlsafe(32)
        from datetime import datetime, timedelta
        reset_tokens[token] = {"email": email, "expira": datetime.utcnow() + timedelta(hours=1)}
        
        asunto = "Recuperacion de contrasena - CarniMarket"
        html = f"""
        <html>
        <body style="font-family: Arial; padding: 20px;">
            <h2 style="color: #c0392b;">CarniMarket</h2>
            <p>Has solicitado recuperar tu contrasena.</p>
            <p>Este enlace expira en 1 hora.</p>
            <p>Tu email de recuperacion: {email}</p>
        </body>
        </html>
        """
        enviar_correo(email, asunto, html)
    
    return {"mensaje": "Si el correo existe, recibiras las instrucciones"}

@app.post("/auth/reset-password")
def cambiar_contrasena(data: dict, response: Response):
    token = data.get("token", "")
    nueva_password = data.get("password", "")
    
    if token not in reset_tokens:
        return {"error": "Token inválido o expirado"}
    
    if len(nueva_password) < 6:
        return {"error": "La contraseña debe tener al menos 6 caracteres"}
    
    global ADMIN_PASSWORD
    ADMIN_PASSWORD = pwd_context.hash(nueva_password)
    del reset_tokens[token]
    
    nuevo_token = crear_token({"sub": ADMIN_USER})
    response.set_cookie(key="token", value=nuevo_token, httponly=True, max_age=28800, samesite="lax", secure=True)
    return {"mensaje": "Contraseña actualizada", "token": nuevo_token}

# --- GESTIÓN DE PRODUCTOS (INVENTARIO) ---

@app.get("/inventario")
def listar_inventario(db: Session = Depends(get_db)):
    productos = db.query(Producto).order_by(Producto.nombre).all()
    result = {}
    for p in productos:
        try:
            tipo_prod = "kilo"
            if hasattr(p, 'tipo') and p.tipo:
                tipo_prod = p.tipo
            result[p.nombre] = {
                "id": p.id,
                "stock": p.stock,
                "minimo": p.minimo,
                "precio_kilo": p.precio_kilo,
                "tipo": tipo_prod
            }
        except:
            pass
    return result

@app.post("/admin/producto")
def crear_producto(p: ProductoRequest, db: Session = Depends(get_db)):
    tipo = p.tipo if p.tipo in ["kilo", "plato"] else "kilo"
    existe = db.query(Producto).filter(Producto.nombre == p.nombre).first()
    if existe:
        return {"error": "El producto ya existe"}
    nuevo = Producto(nombre=p.nombre, stock=p.stock, minimo=p.minimo, precio_kilo=p.precio_kilo, tipo=tipo)
    db.add(nuevo)
    db.commit()
    ahora = obtener_hora_colombia()
    db.add(Historial(producto=p.nombre, tipo="entrada", cantidad=p.stock, motivo="Carga Inicial", fecha=ahora))
    db.commit()
    return {"mensaje": "Producto creado con éxito"}

@app.put("/admin/producto/{id}")
def editar_producto_completo(id: int, data: ProductoUpdate, db: Session = Depends(get_db)):
    p = db.query(Producto).filter(Producto.id == id).first()
    if not p: return {"error": "No existe"}
    
    if data.precio_kilo is not None: p.precio_kilo = data.precio_kilo
    if data.minimo is not None: p.minimo = data.minimo
    if data.nombre is not None: p.nombre = data.nombre
    if data.tipo is not None and data.tipo in ["kilo", "plato"]: p.tipo = data.tipo
    
    if data.stock is not None:
        dif = data.stock - p.stock
        if dif != 0:
            db.add(Historial(producto=p.nombre, tipo="ajuste", cantidad=abs(dif), motivo="Corrección manual", fecha=obtener_hora_colombia()))
        p.stock = data.stock
        
    db.commit()
    return {"mensaje": "Producto actualizado"}

@app.delete("/admin/producto/{id}")
def eliminar_producto_por_id(id: int, db: Session = Depends(get_db)):
    p = db.query(Producto).filter(Producto.id == id).first()
    if p:
        db.delete(p)
        db.commit()
        return {"mensaje": "Producto eliminado"}
    return {"error": "No encontrado"}

# --- GESTIÓN DE CLIENTES ---

@app.get("/admin/clientes")
def get_clientes(db: Session = Depends(get_db)):
    clientes = db.query(Cliente).order_by(Cliente.nombre).all()
    return [{"id": c.id, "nombre": c.nombre, "telefono": c.telefono, "direccion": c.direccion, "fecha_registro": c.fecha_registro.strftime("%Y-%m-%d") if c.fecha_registro else "—"} for c in clientes]

@app.post("/admin/cliente")
def crear_cliente(c: ClienteRequest, db: Session = Depends(get_db)):
    nuevo = Cliente(nombre=c.nombre, telefono=c.telefono, direccion=c.direccion, fecha_registro=obtener_hora_colombia())
    db.add(nuevo)
    db.commit()
    return {"mensaje": "Cliente guardado"}

@app.put("/admin/cliente/{id}")
def editar_cliente_completo(id: int, data: ClienteUpdate, db: Session = Depends(get_db)):
    c = db.query(Cliente).filter(Cliente.id == id).first()
    if not c: return {"error": "No encontrado"}
    if data.nombre is not None: c.nombre = data.nombre
    if data.telefono is not None: c.telefono = data.telefono
    if data.direccion is not None: c.direccion = data.direccion
    db.commit()
    return {"mensaje": "Cliente actualizado"}

@app.delete("/admin/cliente/{id}")
def eliminar_cliente(id: int, db: Session = Depends(get_db)):
    c = db.query(Cliente).filter(Cliente.id == id).first()
    if c:
        db.delete(c)
        db.commit()
        return {"mensaje": "Cliente eliminado"}
    return {"error": "No existe"}

# --- GESTIÓN DE VENTAS ---

@app.post("/vender")
def registrar_venta(v: VentaRequest, db: Session = Depends(get_db)):
    p = db.query(Producto).filter(Producto.nombre == v.producto).first()
    if not p: return {"error": "El producto no existe"}
    
    tipo_producto = getattr(p, 'tipo', 'kilo') or 'kilo'
    es_encargado = v.pagado == "encargado"
    
    # Convertir según el tipo de venta y producto
    if tipo_producto == "plato":
        if v.unidad != "plato":
            return {"error": "Este producto se vende solo por platos"}
        if v.cantidad > p.stock:
            return {"error": f"Stock insuficiente ({int(p.stock)} platos)"}
        # Solo descontar si NO es encargado
        if not es_encargado:
            p.stock -= v.cantidad
        kilos = 0
        total = v.cantidad * p.precio_kilo
    else:
        if v.unidad == "gramos":
            kilos = v.cantidad / 1000
        elif v.unidad == "plato":
            return {"error": "Este producto se vende por kilos/gramos, no por platos"}
        else:
            kilos = v.cantidad
        
        if kilos > p.stock:
            return {"error": f"Stock insuficiente ({p.stock}kg)"}
        
        # Solo descontar si NO es encargado
        if not es_encargado:
            p.stock -= kilos
        total = kilos * p.precio_kilo
    
    ahora = obtener_hora_colombia()
    
    # Procesar fechas
    fecha_venta = ahora
    if v.fecha_venta:
        try:
            fecha_venta = datetime.strptime(v.fecha_venta, "%Y-%m-%d").replace(hour=ahora.hour, minute=ahora.minute, second=ahora.second)
        except:
            fecha_venta = ahora
    
    fecha_vencimiento = None
    if v.fecha_vencimiento:
        try:
            fecha_vencimiento = datetime.strptime(v.fecha_vencimiento, "%Y-%m-%d")
        except:
            fecha_vencimiento = None
    
    nueva_venta = Venta(
        producto=v.producto,
        kilos=kilos,
        precio_kilo=p.precio_kilo,
        subtotal=total,
        fecha_venta=fecha_venta,
        fecha_vencimiento=fecha_vencimiento,
        cliente_nombre=v.cliente_nombre,
        pagado=v.pagado,
        notas=v.notas
    )
    db.add(nueva_venta)
    db.add(Historial(producto=v.producto, tipo="salida", cantidad=kilos, motivo=f"Venta a {v.cliente_nombre}", fecha=ahora))
    db.commit()
    
    if tipo_producto == "plato":
        return {"mensaje": "Venta exitosa", "producto": v.producto, "platos": int(v.cantidad), "subtotal": total, "stock_restante": int(p.stock)}
    else:
        return {"mensaje": "Venta exitosa", "producto": v.producto, "kilos": kilos, "subtotal": total, "stock_restante": p.stock}

@app.get("/admin/ventas")
def listar_ventas(
    fecha_inicio: Optional[str] = None,
    fecha_fin: Optional[str] = None,
    db: Session = Depends(get_db)):
    query = db.query(Venta)
    
    if fecha_inicio:
        fi = datetime.strptime(fecha_inicio, "%Y-%m-%d").replace(hour=0, minute=0, second=0)
        query = query.filter(Venta.fecha_venta >= fi)
    if fecha_fin:
        ff = datetime.strptime(fecha_fin, "%Y-%m-%d").replace(hour=23, minute=59, second=59)
        query = query.filter(Venta.fecha_venta <= ff)
    
    ventas = query.order_by(Venta.fecha_venta.desc()).limit(150).all()
    return [{"id": v.id, "fecha_venta": v.fecha_venta.strftime("%Y-%m-%d %H:%M"), "cliente": v.cliente_nombre, "producto": v.producto, "kilos": v.kilos, "subtotal": v.subtotal, "pagado": v.pagado, "notas": v.notas, "fecha_vencimiento": v.fecha_vencimiento.strftime("%Y-%m-%d") if v.fecha_vencimiento else ""} for v in ventas]

@app.get("/admin/encargados")
def listar_encargados(db: Session = Depends(get_db)):
    ventas = db.query(Venta).filter(Venta.pagado == "encargado").order_by(Venta.fecha_venta.desc()).all()
    return [{"id": v.id, "fecha_venta": v.fecha_venta.strftime("%Y-%m-%d %H:%M"), "cliente": v.cliente_nombre, "producto": v.producto, "kilos": v.kilos, "subtotal": v.subtotal, "notas": v.notas} for v in ventas]

@app.get("/admin/encargados/toggle/{id}")
def toggle_encargado(id: int, db: Session = Depends(get_db)):
    v = db.query(Venta).filter(Venta.id == id).first()
    if not v: return {"error": "No existe"}
    
    p = db.query(Producto).filter(Producto.nombre == v.producto).first()
    
    # Si cambia de encargado a pagado/debe: descontar stock
    if v.pagado == "encargado" and p:
        p.stock -= v.kilos
    # Si cambia de pagado/debe a encargado: restaurar stock
    elif v.pagado != "encargado" and p:
        p.stock += v.kilos
    
    # Nuevo estado
    if v.pagado == "encargado":
        v.pagado = "pagado"
    elif v.pagado == "pagado":
        v.pagado = "debe"
    else:
        v.pagado = "encargado"
    
    db.commit()
    return {"mensaje": "Estado actualizado", "pagado": v.pagado}

@app.put("/admin/encargados/{id}")
def confirmar_encargado(id: int, data: dict, db: Session = Depends(get_db)):
    v = db.query(Venta).filter(Venta.id == id).first()
    if not v: return {"error": "No existe"}
    
    nuevo_estado = data.get("estado", "pagado")
    if nuevo_estado not in ["pagado", "debe"]:
        return {"error": "Estado inválido"}
    
    p = db.query(Producto).filter(Producto.nombre == v.producto).first()
    
    # Solo descontar stock si venia de encargado
    if v.pagado == "encargado" and p and v.kilos:
        p.stock -= v.kilos
    
    v.pagado = nuevo_estado
    db.commit()
    return {"mensaje": "Pedido confirmado", "pagado": v.pagado}
    
@app.put("/admin/venta/{id}")
def corregir_venta(id: int, data: dict, db: Session = Depends(get_db), token: str = Cookie(default=None)):
    if not verificar_token(token):
        raise HTTPException(status_code=401)
    u = db.query(Usuario).filter(Usuario.email == verificar_token(token), Usuario.activo == True).first()
    if not u or u.rol not in ["admin", "dueno", "empleado"]:
        raise HTTPException(status_code=403)
    
    v = db.query(Venta).filter(Venta.id == id).first()
    if not v: return {"error": "No existe"}
    if "pagado" in data: v.pagado = data["pagado"]
    if "notas" in data: v.notas = data["notas"]
    if "kilos" in data:
        dif = data["kilos"] - v.kilos
        p = db.query(Producto).filter(Producto.nombre == v.producto).first()
        if p: p.stock -= dif
        v.kilos = data["kilos"]
        v.subtotal = v.kilos * v.precio_kilo
    db.commit()
    return {"mensaje": "Venta corregida"}

@app.put("/admin/venta/{id}/pago")
def toggle_pago_venta(id: int, db: Session = Depends(get_db)):
    v = db.query(Venta).filter(Venta.id == id).first()
    if not v: return {"error": "No existe"}
    #轮回切换: encargado -> pagado -> debe -> encargado
    if v.pagado == "encargado":
        v.pagado = "pagado"
    elif v.pagado == "pagado":
        v.pagado = "debe"
    else:
        v.pagado = "encargado"
    db.commit()
    return {"mensaje": "Estado de pago actualizado", "pagado": v.pagado}

@app.delete("/admin/venta/{id}")
def eliminar_venta(id: int, db: Session = Depends(get_db)):
    v = db.query(Venta).filter(Venta.id == id).first()
    if not v: return {"error": "No existe"}
    # Restaurar stock del producto
    p = db.query(Producto).filter(Producto.nombre == v.producto).first()
    if p:
        p.stock += v.kilos
    db.delete(v)
    db.commit()
    return {"mensaje": "Venta eliminada"}

# --- GESTIÓN DE GASTOS ---

@app.post("/admin/gasto")
def registrar_gasto(g: GastoRequest, db: Session = Depends(get_db)):
    nuevo = Gasto(descripcion=g.descripcion, categoria=g.categoria, monto=g.monto, fecha=obtener_hora_colombia())
    db.add(nuevo)
    db.commit()
    return {"mensaje": "Gasto registrado"}

@app.get("/admin/gastos")
def listar_gastos(db: Session = Depends(get_db)):
    gs = db.query(Gasto).order_by(Gasto.fecha.desc()).all()
    return [{"id": g.id, "fecha": g.fecha.strftime("%Y-%m-%d"), "descripcion": g.descripcion, "monto": g.monto, "categoria": g.categoria} for g in gs]

@app.delete("/admin/gasto/{id}")
def eliminar_gasto(id: int, db: Session = Depends(get_db)):
    g = db.query(Gasto).filter(Gasto.id == id).first()
    if not g: return {"error": "No existe"}
    db.delete(g)
    db.commit()
    return {"mensaje": "Gasto eliminado"}

# --- REPORTES Y CAJA ---

@app.get("/admin/caja")
def estado_caja(db: Session = Depends(get_db)):
    ingresos = db.query(func.sum(Venta.subtotal)).filter(Venta.pagado == "pagado").scalar() or 0
    egresos = db.query(func.sum(Gasto.monto)).scalar() or 0
    return {"saldo_real": ingresos - egresos, "ingresos": ingresos, "egresos": egresos}

@app.get("/admin/caja-detalle")
def caja_detalle(
    fecha_inicio: Optional[str] = None,
    fecha_fin: Optional[str] = None,
    db: Session = Depends(get_db)):
    try:
        query_vp = db.query(Venta).filter(Venta.pagado == "pagado")
        query_vd = db.query(Venta).filter(Venta.pagado == "debe")
        query_g = db.query(Gasto)
        
        if fecha_inicio:
            fi = datetime.strptime(fecha_inicio, "%Y-%m-%d").replace(hour=0, minute=0, second=0)
            query_vp = query_vp.filter(Venta.fecha_venta >= fi)
            query_vd = query_vd.filter(Venta.fecha_venta >= fi)
            query_g = query_g.filter(Gasto.fecha >= fi)
        
        if fecha_fin:
            ff = datetime.strptime(fecha_fin, "%Y-%m-%d").replace(hour=23, minute=59, second=59)
            query_vp = query_vp.filter(Venta.fecha_venta <= ff)
            query_vd = query_vd.filter(Venta.fecha_venta <= ff)
            query_g = query_g.filter(Gasto.fecha <= ff)
        
        vp = sum(v.subtotal for v in query_vp.all())
        vd = sum(v.subtotal for v in query_vd.all())
        g = sum(g_.monto for g_ in query_g.all())
        
        return {
            "ventas_pagadas": vp,
            "ventas_deben": vd,
            "total_ventas": vp + vd,
            "gastos": g,
            "saldo_real": vp - g
        }
    except Exception as e:
        return {"error": str(e), "ventas_pagadas": 0, "ventas_deben": 0, "total_ventas": 0, "gastos": 0, "saldo_real": 0}

@app.get("/admin/dashboard")
def get_dashboard_data(periodo: str = "7dias", db: Session = Depends(get_db)):
    ahora = obtener_hora_colombia()
    inicio_dia = ahora.replace(hour=0, minute=0, second=0, microsecond=0)
    fin_dia = ahora.replace(hour=23, minute=59, second=59, microsecond=999999)
    
    if periodo == "hoy":
        fecha_inicio = inicio_dia
    elif periodo == "7dias":
        fecha_inicio = inicio_dia - timedelta(days=6)
    elif periodo == "30dias":
        fecha_inicio = ahora.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    elif periodo == "todo":
        fecha_inicio = datetime(2020, 1, 1)
    else:
        fecha_inicio = inicio_dia - timedelta(days=6)
    
    v_periodo = db.query(Venta).filter(Venta.fecha_venta >= fecha_inicio, Venta.fecha_venta <= fin_dia).all()
    
    conteo = {}
    for v in v_periodo:
        if v.producto:
            conteo[v.producto] = conteo.get(v.producto, 0) + (v.kilos or 0)
    
    alertas = []
    try:
        prods = db.query(Producto).filter(Producto.stock <= Producto.minimo).all()
        alertas = [{"nombre": p.nombre, "stock": p.stock} for p in prods]
    except:
        pass
    
    inicio_mes = ahora.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    v_mes = db.query(Venta).filter(Venta.fecha_venta >= inicio_mes).all()
    
    return {
        "total_hoy": sum(v.subtotal or 0 for v in db.query(Venta).filter(Venta.fecha_venta >= inicio_dia, Venta.fecha_venta <= fin_dia).all()),
        "total_periodo": sum(v.subtotal or 0 for v in v_periodo),
        "total_mes": sum(v.subtotal or 0 for v in v_mes),
        "productos_ventas": dict(sorted(conteo.items(), key=lambda x: x[1], reverse=True)[:7]),
        "stock_bajo": alertas,
        "periodo": periodo
    }

@app.get("/admin/deudas")
def get_reporte_deudas(db: Session = Depends(get_db), fecha_inicio: Optional[str] = None, fecha_fin: Optional[str] = None):
    query = db.query(Venta).filter(Venta.pagado == "debe")

    if fecha_inicio:
        try:
            fecha_inicio_dt = datetime.strptime(fecha_inicio, "%Y-%m-%d")
            query = query.filter(Venta.fecha_venta >= fecha_inicio_dt)
        except ValueError:
            return {"error": "Formato de fecha_inicio inválido. Use YYYY-MM-DD."}

    if fecha_fin:
        try:
            fecha_fin_dt = datetime.strptime(fecha_fin, "%Y-%m-%d")
            query = query.filter(Venta.fecha_venta <= fecha_fin_dt)
        except ValueError:
            return {"error": "Formato de fecha_fin inválido. Use YYYY-MM-DD."}

    pendientes = query.all()
    resumen = {}
    for v in pendientes:
        if v.cliente_nombre not in resumen:
            cli = db.query(Cliente).filter(Cliente.nombre == v.cliente_nombre).first()
            resumen[v.cliente_nombre] = {
                "total": 0,
                "tel": cli.telefono if cli else "",
                "direccion": cli.direccion if cli else "",
                "fecha_vencimiento": None
            }
        resumen[v.cliente_nombre]["total"] += v.subtotal
        if v.fecha_vencimiento:
            if resumen[v.cliente_nombre]["fecha_vencimiento"] is None:
                resumen[v.cliente_nombre]["fecha_vencimiento"] = v.fecha_vencimiento
            elif v.fecha_vencimiento < resumen[v.cliente_nombre]["fecha_vencimiento"]:
                resumen[v.cliente_nombre]["fecha_vencimiento"] = v.fecha_vencimiento

    res = []
    for n, d in resumen.items():
        fecha_venc_str = d["fecha_vencimiento"].strftime("%Y-%m-%d") if d["fecha_vencimiento"] else "Sin vencimiento"
        res.append({
            "cliente": n,
            "total": d["total"],
            "direccion": d["direccion"],
            "fecha_vencimiento": fecha_venc_str,
            "whatsapp_link": f"https://wa.me/57{d['tel']}?text="+urllib.parse.quote(f"Hola {n}, recuerda que tienes un saldo pendiente con Carnimarket de : ${d['total']:,.0f}") if d['tel'] else ""
        })
    return {"deudas": res, "total_pendiente": sum(x["total"] for x in res)}

@app.get("/admin/historial")
def ver_historial_movimientos(db: Session = Depends(get_db)):
    logs = db.query(Historial).order_by(Historial.fecha.desc()).limit(100).all()
    return [{"fecha": l.fecha.strftime("%Y-%m-%d %H:%M"), "producto": l.producto, "tipo": l.tipo, "cantidad": l.cantidad, "motivo": l.motivo} for l in logs]

@app.get("/admin/exportar/caja")
def generar_excel_caja(
    fecha_inicio: Optional[str] = None,
    fecha_fin: Optional[str] = None,
    db: Session = Depends(get_db)):
    query_v = db.query(Venta)
    query_g = db.query(Gasto)
    
    if fecha_inicio:
        fi = datetime.strptime(fecha_inicio, "%Y-%m-%d").replace(hour=0, minute=0, second=0)
        query_v = query_v.filter(Venta.fecha_venta >= fi)
        query_g = query_g.filter(Gasto.fecha >= fi)
    if fecha_fin:
        ff = datetime.strptime(fecha_fin, "%Y-%m-%d").replace(hour=23, minute=59, second=59)
        query_v = query_v.filter(Venta.fecha_venta <= ff)
        query_g = query_g.filter(Gasto.fecha <= ff)
    
    filename = "Caja"
    if fecha_inicio or fecha_fin:
        filename += f"_{fecha_inicio or 'inicio'}_{fecha_fin or 'fin'}"
    filename += ".xlsx"
    
    wb = Workbook()
    ws_v = wb.create_sheet("Ventas")
    ws_v.append(["ID", "Fecha", "Cliente", "Producto", "Kilos", "Precio/Kg", "Total", "Estado"])
    for v in query_v.order_by(Venta.fecha_venta.desc()).all():
        ws_v.append([v.id, v.fecha_venta.strftime("%Y-%m-%d %H:%M"), v.cliente_nombre, v.producto, v.kilos, v.precio_kilo, v.subtotal, v.pagado])
    
    ws_g = wb.create_sheet("Gastos")
    ws_g.append(["ID", "Fecha", "Descripción", "Categoría", "Monto"])
    for g in query_g.order_by(Gasto.fecha.desc()).all():
        ws_g.append([g.id, g.fecha.strftime("%Y-%m-%d %H:%M"), g.descripcion, g.categoria, g.monto])
    
    stream = io.BytesIO(); wb.save(stream); stream.seek(0)
    return StreamingResponse(stream, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": f"attachment; filename={filename}"})

@app.get("/admin/exportar/ventas")
def generar_excel_ventas(
    fecha_inicio: Optional[str] = None,
    fecha_fin: Optional[str] = None,
    db: Session = Depends(get_db)):
    query = db.query(Venta)
    
    if fecha_inicio:
        fi = datetime.strptime(fecha_inicio, "%Y-%m-%d").replace(hour=0, minute=0, second=0)
        query = query.filter(Venta.fecha_venta >= fi)
    if fecha_fin:
        ff = datetime.strptime(fecha_fin, "%Y-%m-%d").replace(hour=23, minute=59, second=59)
        query = query.filter(Venta.fecha_venta <= ff)
    
    ventas = query.order_by(Venta.fecha_venta.desc()).all()
    filename = "Ventas"
    if fecha_inicio or fecha_fin:
        filename += f"_{fecha_inicio or 'inicio'}_{fecha_fin or 'fin'}"
    filename += ".xlsx"
    
    wb = Workbook(); ws = wb.active; ws.title = "Reporte"
    headers = ["ID", "Fecha", "Cliente", "Producto", "Kilos", "Precio/Kg", "Total", "Estado", "Vencimiento"]
    ws.append(headers)
    for v in ventas:
        ws.append([v.id, v.fecha_venta.strftime("%Y-%m-%d %H:%M"), v.cliente_nombre, v.producto, v.kilos, v.precio_kilo, v.subtotal, v.pagado, v.fecha_vencimiento.strftime("%Y-%m-%d") if v.fecha_vencimiento else ""])
    stream = io.BytesIO(); wb.save(stream); stream.seek(0)
    return StreamingResponse(stream, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": f"attachment; filename={filename}"})